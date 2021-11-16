import csv
import random
import re
import os
from flask import current_app

import chardet
import sys
from openpyxl import load_workbook
from unidecode import unidecode

import logging

logging.basicConfig(level=logging.DEBUG, format='[%(levelname)s] %(asctime)s - %(message)s',
                    datefmt='%d-%b-%y %H:%M:%S')


def get_file_encoding(src_file_path):
    """
    Get the encoding type of a file
    :param src_file_path: file path
    :return: str - file encoding type
    """
    with open(src_file_path) as src_file:
        return src_file.encoding


def get_file_encoding_chardet(file_path):
    """
    Get the encoding of a file using chardet package
    :param file_path:
    :return:
    """
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
        return result['encoding']


def compose_username(f_name, l_name, etablissment, u_names):
    f_name = re.sub("[',\"]", '', f_name.lower())
    l_name = re.sub("[',\"]", '', l_name.lower())
    etablissment = re.sub("[',\"]", '', etablissment.lower())
    gen = unidecode("{}_{}".format(
        f_name.split()[0].replace('-', '_'),
        "_".join(etablissment.split()[0:2]).replace('-', '_').replace('.', '_')
    ))
    copy = gen
    counter = 1
    # eliminate duplicate username(s)
    while gen in u_names:
        gen = "{}_{}".format(copy, counter)
        counter += 1

    return gen


def compose_last_name(f_name):
    split = [name.capitalize() for name in f_name.split()]
    if len(split) > 1:
        return " ".join(split[0:1]), " ".join(split[1:])
    else:
        return split[0], split[0]


def rand_str(count, allowed=None):
    if allowed is None:
        allowed = [chr(x) for x in range(ord('a'), ord('z') + 1)]
    return ''.join(random.choice(allowed) for x in range(count))


def load_usernames(path):
    """
    @todo Load usernames from local cache file (if necessary)
    :param path: Path to root folder containing sub folders
    :return: List already used usernames (from final spreadsheets)
    """
    import glob
    import os
    usernames = []
    head, _ = os.path.split(path)
    files = glob.glob(
        os.path.join(current_app.config['UPLOAD_FOLDER'], "*/{}/*.xlsx".format(current_app.config['GENERATED_DIR'])))
    logging.info("Loading used usernames...")
    for final in files:
        wb = load_workbook(final)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            labels = [[field.value.lower() for field in row if field.value] for row in ws.iter_rows(max_row=1)][0]
            USERNAME_COL = labels.index('username') + 1
            current_row = 1
            for row in ws.iter_rows():
                col = 1
                for cell in row:
                    if cell.value in labels:
                        continue
                    if col == USERNAME_COL:
                        usernames.append(ws.cell(row=current_row, column=USERNAME_COL).value)
                    col += 1
                current_row += 1

    return usernames


def main(config=None):
    if config is None:
        return

    log_file = None
    old_log = sys.stdout
    stats = {"males": 0, "females": 0}
    try:
        log_file = open(config['LOG_FILE'], "w")
        sys.stdout = log_file
        wb = load_workbook(config['FILENAME'])
        usernames = load_usernames(config['SAVE_PATH'])
        pltform = config['PLATFORM']

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            labels = [[field.value.lower() for field in row if field.value] for row in ws.iter_rows(max_row=1)][0]
            # print("WS: ", labels)

            USERNAME_COL = labels.index('username') + 1
            F_NAME_COL = labels.index('firstname') + 1
            L_NAME_COL = labels.index('lastname') + 1
            PASS_COL = labels.index('password') + 1

            PLATFORM_ETABLISSEMENT_TITLE = 'profile_field_' + (
                'etablissement' if pltform == "MON_ECOLE_ENLIGNE" else 'school_name')
            ETABLISSMENT_COL = labels.index(PLATFORM_ETABLISSEMENT_TITLE) + 1

            PLATFORM_GENDER_TITLE = 'profile_field_' + ('sexe' if pltform == "MON_ECOLE_ENLIGNE" else 'sex')
            GENDER_COL = labels.index(PLATFORM_GENDER_TITLE) + 1

            current_row = 1
            for row in ws.iter_rows():
                col = 1
                for cell in row:
                    if cell.value in labels:
                        break
                    if col == USERNAME_COL:  # Stop at Username Column
                        first_name = ws.cell(row=current_row, column=F_NAME_COL).value
                        last_name = ws.cell(row=current_row, column=L_NAME_COL).value
                        etabliss = ws.cell(row=current_row, column=ETABLISSMENT_COL).value
                        password = ws.cell(row=current_row, column=PASS_COL).value
                        if etabliss is None:
                            continue
                        else:
                            etabliss = re.sub(' +', ' ', etabliss)

                        if first_name is None and last_name is None:
                            continue
                        else:
                            if first_name is not None:
                                first_name = re.sub(' +', ' ', first_name)
                            if last_name is not None:
                                last_name = re.sub(' +', ' ', last_name)

                        if config['UPDATE_NAMES']:
                            if last_name is None:
                                first_name, last_name = compose_last_name(first_name)
                                ws.cell(row=current_row, column=F_NAME_COL, value=first_name)
                                ws.cell(row=current_row, column=L_NAME_COL, value=last_name)
                            if first_name is None:
                                first_name, last_name = compose_last_name(last_name)
                                ws.cell(row=current_row, column=F_NAME_COL, value=first_name)
                                ws.cell(row=current_row, column=L_NAME_COL, value=last_name)
                        if config['UPDATE_PASSWORD'] or type(password) == int:
                            # if password is None or type(password) == int:
                            pass_gen = rand_str(config['PASSWORD_LENGTH'])
                            ws.cell(row=current_row, column=PASS_COL, value=pass_gen)

                        if config['UPDATE_USERNAME']:
                            username = compose_username(first_name, last_name, etabliss, usernames)

                            usernames.append(username)
                            ws.cell(row=cell.row, column=cell.column, value=username)
                        gender = ws.cell(row=current_row, column=GENDER_COL).value
                        if len(gender):
                            gen_index = "males" if gender.lower().startswith("m") else "females"
                            stats[gen_index] += 1
                    col += 1
                current_row += 1
            if config['EXPORT_CSV']:
                CSV = "{}/{}.GEN_{}.csv".format(config['SAVE_PATH'], wb.sheetnames.index(sheet_name) + 1,
                                                ws.title.replace(" ", "_"))
                with open(CSV, 'w', newline="", encoding='utf-8') as fh:
                    c = csv.writer(fh, delimiter=config['CSV_DELIMITER'])
                    for r in ws.rows:
                        if r[0].value is not None:  # ignore empty fields.
                            c.writerow([cell.value for cell in r])
                print('Endcoding: ' + str(get_file_encoding(CSV)))
                print('Endcoding with chardet: ' + str(get_file_encoding_chardet(CSV)))
        wb.save("{}/{}.xlsx".format(config['SAVE_PATH'], 'final'))

        print("ALL DONE!!!")

    except Exception as e:
        print(str(e))
    finally:
        if log_file is not None:
            log_file.close()
        sys.stdout = old_log

    return stats
